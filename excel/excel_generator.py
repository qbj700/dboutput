"""
Excel 명세서 생성기
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os
from datetime import datetime


class DBSpecExcelGenerator:
    """DB 산출물 Excel 생성 클래스"""
    
    def __init__(self):
        self.workbook = None
        self.styles = self._create_styles()
        
    def _create_styles(self):
        """Excel 스타일 정의 (이미지 양식에 맞춤)"""
        return {
            # 테이블 정보 라벨 스타일 (회색 배경)
            'info_label': {
                'font': Font(name='맑은 고딕', size=11, bold=True),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'fill': PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            # 테이블 정보 값 스타일 (흰색 배경)
            'info_value': {
                'font': Font(name='맑은 고딕', size=11),
                'alignment': Alignment(horizontal='left', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            # 컬럼 헤더 스타일 (회색 배경)
            'header': {
                'font': Font(name='맑은 고딕', size=11, bold=True),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'fill': PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            # 데이터 스타일 (왼쪽 정렬)
            'data': {
                'font': Font(name='맑은 고딕', size=10),
                'alignment': Alignment(horizontal='left', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            # 데이터 스타일 (가운데 정렬)
            'data_center': {
                'font': Font(name='맑은 고딕', size=10),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            # Primary Key 스타일 (일반 텍스트와 동일)
            'pk': {
                'font': Font(name='맑은 고딕', size=10, bold=False, color='000000'),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            }
        }
        
    def generate_excel(self, metadata, save_path):
        """
        메타데이터를 기반으로 Excel 명세서 생성
        
        Args:
            metadata (dict): 데이터베이스 메타데이터
            save_path (str): 저장할 파일 경로
            
        Returns:
            str: 생성된 파일 경로
        """
        self.workbook = Workbook()
        
        # 기본 시트 제거
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
            
        # 포맷된 메타데이터 준비
        formatted_metadata = self._format_metadata_for_excel(metadata)
        
        # 하나의 테이블명세서 시트에 모든 테이블 나열
        self._create_unified_table_sheet(formatted_metadata)
            
        # 파일 저장
        save_dir = os.path.dirname(save_path)
        if save_dir:  # 디렉토리가 있는 경우에만 생성
            os.makedirs(save_dir, exist_ok=True)
        self.workbook.save(save_path)
        
        return save_path
    
    def generate_table_list_excel(self, table_list_data, save_path):
        """
        테이블 목록 Excel 생성
        
        Args:
            table_list_data (dict): 테이블 목록 데이터
            save_path (str): 저장할 파일 경로
            
        Returns:
            str: 생성된 파일 경로
        """
        self.workbook = Workbook()
        
        # 기본 시트 제거
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
            
        # 테이블 목록 시트 생성
        self._create_table_list_sheet(table_list_data)
            
        # 파일 저장
        save_dir = os.path.dirname(save_path)
        if save_dir:  # 디렉토리가 있는 경우에만 생성
            os.makedirs(save_dir, exist_ok=True)
        self.workbook.save(save_path)
        
        return save_path
    
    def _create_table_list_sheet(self, table_list_data):
        """테이블 목록 시트 생성"""
        ws = self.workbook.create_sheet("테이블목록")
        
        # 헤더 생성
        headers = ['NO', '테이블명', '논리명']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_style(cell, self.styles['header'])
        
        # 테이블 목록 데이터 추가
        for idx, table_info in enumerate(table_list_data['table_list'], 2):
            # NO
            cell = ws.cell(row=idx, column=1, value=table_info['no'])
            self._apply_style(cell, self.styles['data_center'])
            
            # 테이블명
            cell = ws.cell(row=idx, column=2, value=table_info['table_name'])
            self._apply_style(cell, self.styles['data'])
            
            # 논리명 (테이블 코멘트)
            cell = ws.cell(row=idx, column=3, value=table_info['table_comment'])
            self._apply_style(cell, self.styles['data'])
        
        # 컬럼 너비 조정
        column_widths = [8, 30, 50]  # NO, 테이블명, 논리명
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    def _create_unified_table_sheet(self, formatted_metadata):
        """모든 테이블을 하나의 시트에 통합 생성"""
        ws = self.workbook.create_sheet("테이블명세서")
        
        current_row = 1
        table_count = 0
        
        # 각 테이블을 순차적으로 배치
        for table_name, table_data in formatted_metadata['tables_by_name'].items():
            # 첫 번째 테이블이 아니면 간격 추가 (2행 띄어서)
            if table_count > 0:
                current_row += 2
                
            current_row = self._add_table_to_sheet(ws, table_name, table_data, formatted_metadata['foreign_keys'], formatted_metadata['indexes'], current_row)
            table_count += 1
    
    def _add_table_to_sheet(self, ws, table_name, table_data, foreign_keys, indexes, start_row):
        """시트에 테이블 정보를 추가 (기존 _create_table_sheet 로직 재사용)"""
        
        # 테이블 정보 영역 (시작행부터 7행)
        
        # 1행: 테이블명
        ws.cell(row=start_row, column=1, value="테이블명")
        ws.cell(row=start_row, column=3, value=table_name)
        ws.merge_cells(f'A{start_row}:B{start_row}')
        ws.merge_cells(f'C{start_row}:J{start_row}')
        self._apply_style(ws.cell(row=start_row, column=1), self.styles['info_label'])
        self._apply_style(ws.cell(row=start_row, column=2), self.styles['info_label'])
        self._apply_style(ws.cell(row=start_row, column=3), self.styles['info_value'])
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J']:
            col_num = ord(col) - ord('A') + 1
            self._apply_style(ws.cell(row=start_row, column=col_num), self.styles['info_value'])
        
        # 2행: 논리명 (테이블 comment)
        row2 = start_row + 1
        ws.cell(row=row2, column=1, value="논리명")
        ws.cell(row=row2, column=3, value=table_data['table_info']['comment'] or '')
        ws.merge_cells(f'A{row2}:B{row2}')
        ws.merge_cells(f'C{row2}:J{row2}')
        self._apply_style(ws.cell(row=row2, column=1), self.styles['info_label'])
        self._apply_style(ws.cell(row=row2, column=2), self.styles['info_label'])
        self._apply_style(ws.cell(row=row2, column=3), self.styles['info_value'])
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J']:
            col_num = ord(col) - ord('A') + 1
            self._apply_style(ws.cell(row=row2, column=col_num), self.styles['info_value'])
        
        # 3행: 테이블 설명 (빈 칸)
        row3 = start_row + 2
        ws.cell(row=row3, column=1, value="테이블 설명")
        ws.merge_cells(f'A{row3}:B{row3}')
        ws.merge_cells(f'C{row3}:J{row3}')
        self._apply_style(ws.cell(row=row3, column=1), self.styles['info_label'])
        self._apply_style(ws.cell(row=row3, column=2), self.styles['info_label'])
        self._apply_style(ws.cell(row=row3, column=3), self.styles['info_value'])
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J']:
            col_num = ord(col) - ord('A') + 1
            self._apply_style(ws.cell(row=row3, column=col_num), self.styles['info_value'])
        
        # Primary Key, Foreign Key, Index 정보 수집
        primary_keys = []
        foreign_key_info = []
        
        for col_info in table_data['columns']:
            if col_info['key'] == 'PRI':
                primary_keys.append(col_info['name'])
                
        for fk in foreign_keys:
            if fk['table_name'] == table_name:
                foreign_key_info.append(f"{fk['column_name']}")
        
        # 4행: PRIMARY KEY
        row4 = start_row + 3
        ws.cell(row=row4, column=1, value="PRIMARY KEY")
        ws.cell(row=row4, column=3, value=', '.join(primary_keys) if primary_keys else '')
        ws.merge_cells(f'A{row4}:B{row4}')
        ws.merge_cells(f'C{row4}:J{row4}')
        self._apply_style(ws.cell(row=row4, column=1), self.styles['info_label'])
        self._apply_style(ws.cell(row=row4, column=2), self.styles['info_label'])
        self._apply_style(ws.cell(row=row4, column=3), self.styles['info_value'])
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J']:
            col_num = ord(col) - ord('A') + 1
            self._apply_style(ws.cell(row=row4, column=col_num), self.styles['info_value'])
        
        # 5행: FOREIGN KEY
        row5 = start_row + 4
        ws.cell(row=row5, column=1, value="FOREIGN KEY")
        ws.cell(row=row5, column=3, value=', '.join(foreign_key_info) if foreign_key_info else '')
        ws.merge_cells(f'A{row5}:B{row5}')
        ws.merge_cells(f'C{row5}:J{row5}')
        self._apply_style(ws.cell(row=row5, column=1), self.styles['info_label'])
        self._apply_style(ws.cell(row=row5, column=2), self.styles['info_label'])
        self._apply_style(ws.cell(row=row5, column=3), self.styles['info_value'])
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J']:
            col_num = ord(col) - ord('A') + 1
            self._apply_style(ws.cell(row=row5, column=col_num), self.styles['info_value'])
        
        # 인덱스 정보 수집 및 구성
        table_indexes = [idx for idx in indexes if idx['table_name'] == table_name]
        regular_indexes = []
        unique_indexes = []
        
        # 인덱스를 그룹화하여 "인덱스명(컬럼명)" 형태로 구성
        index_groups = {}
        for idx in table_indexes:
            index_name = idx['index_name']
            column_name = idx['column_name']
            is_unique = idx['non_unique'] == 0 or idx['non_unique'] == False
            
            if index_name not in index_groups:
                index_groups[index_name] = {
                    'columns': [],
                    'is_unique': is_unique
                }
            
            index_groups[index_name]['columns'].append(column_name)
        
        # 인덱스명(컬럼명) 형태로 변환
        for index_name, info in index_groups.items():
            columns_str = ','.join(info['columns'])
            index_display = f"{index_name}({columns_str})"
            
            if info['is_unique']:
                unique_indexes.append(index_display)
            else:
                regular_indexes.append(index_display)
        
        # 6행: INDEX
        row6 = start_row + 5
        ws.cell(row=row6, column=1, value="INDEX")
        ws.cell(row=row6, column=3, value=', '.join(regular_indexes) if regular_indexes else '')
        ws.merge_cells(f'A{row6}:B{row6}')
        ws.merge_cells(f'C{row6}:J{row6}')
        self._apply_style(ws.cell(row=row6, column=1), self.styles['info_label'])
        self._apply_style(ws.cell(row=row6, column=2), self.styles['info_label'])
        self._apply_style(ws.cell(row=row6, column=3), self.styles['info_value'])
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J']:
            col_num = ord(col) - ord('A') + 1
            self._apply_style(ws.cell(row=row6, column=col_num), self.styles['info_value'])
        
        # 7행: UNIQUE INDEX
        row7 = start_row + 6
        ws.cell(row=row7, column=1, value="UNIQUE INDEX")
        ws.cell(row=row7, column=3, value=', '.join(unique_indexes) if unique_indexes else '')
        ws.merge_cells(f'A{row7}:B{row7}')
        ws.merge_cells(f'C{row7}:J{row7}')
        self._apply_style(ws.cell(row=row7, column=1), self.styles['info_label'])
        self._apply_style(ws.cell(row=row7, column=2), self.styles['info_label'])
        self._apply_style(ws.cell(row=row7, column=3), self.styles['info_value'])
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J']:
            col_num = ord(col) - ord('A') + 1
            self._apply_style(ws.cell(row=row7, column=col_num), self.styles['info_value'])
        
        # 8행: 컬럼 헤더
        header_row = start_row + 7
        headers = ['NO', 'PK', 'AI', 'FK', 'NULL', '컬럼명', 'TYPE', 'DEFAULT', '설명', '참조 테이블']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            self._apply_style(cell, self.styles['header'])
            
        # 컬럼 정보 추가 (순차적 NO 부여)
        current_data_row = header_row
        for idx, col_info in enumerate(table_data['columns'], 1):
            current_data_row += 1
            
            # 외래키 참조 정보 찾기 및 FK 여부 판별
            fk_ref = ""
            is_foreign_key = False
            for fk in foreign_keys:
                if (fk['table_name'] == table_name and 
                    fk['column_name'] == col_info['name']):
                    fk_ref = f"{fk['referenced_table_name']}.{fk['referenced_column_name']}"
                    is_foreign_key = True
                    break
                    
            # 데이터 입력
            data_row = [
                idx,  # NO (순차적)
                'Y' if col_info['key'] == 'PRI' else '',  # PK
                'Y' if 'auto_increment' in str(col_info['extra']).lower() else '',  # AI
                'Y' if is_foreign_key else '',  # FK (외래키 정보로 정확히 판별)
                'Y' if col_info['nullable'] in ['YES', 'Y', 1, '1'] else '',  # NULL
                col_info['name'],  # 컬럼명
                col_info['type'],  # TYPE
                col_info['default'] or '',  # DEFAULT
                col_info['comment'] or '',  # 설명
                fk_ref  # 참조 테이블
            ]
            
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=current_data_row, column=col, value=value)
                
                # 스타일 적용
                if col in [1, 2, 3, 4, 5]:  # NO, PK, AI, FK, NULL
                    style = self.styles['data_center']
                    # PK가 Y인 경우 빨간색으로 강조
                    if col == 2 and value == 'Y':
                        style = self.styles['pk']
                else:
                    style = self.styles['data']
                    
                self._apply_style(cell, style)
        
        # 컬럼 너비 조정 (첫 번째 테이블일 때만)
        if start_row == 1:
            column_widths = [12, 4, 6, 6, 8, 20, 20, 15, 30, 20]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
                
        # 다음 테이블 시작 위치 반환
        return current_data_row
        
    def _create_overview_sheet(self, metadata):
        """개요 시트 생성"""
        ws = self.workbook.create_sheet("📊 데이터베이스 개요", 0)
        
        # 제목
        ws['A1'] = "데이터베이스 명세서"
        ws.merge_cells('A1:F1')
        self._apply_style(ws['A1'], self.styles['title'])
        
        # 기본 정보
        info = metadata['connection_info']
        
        row = 3
        info_items = [
            ("데이터베이스 종류:", info['dbms']),
            ("서버 주소:", f"{info['host']}:{info['port']}"),
            ("데이터베이스명:", info['database']),
            ("버전:", info['version']),
            ("생성일시:", info['collection_time'])
        ]
        
        for label, value in info_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            self._apply_style(ws[f'A{row}'], {
                'font': Font(name='맑은 고딕', size=11, bold=True),
                'alignment': Alignment(horizontal='right', vertical='center')
            })
            self._apply_style(ws[f'B{row}'], {
                'font': Font(name='맑은 고딕', size=11),
                'alignment': Alignment(horizontal='left', vertical='center')
            })
            row += 1
            
        # 통계 정보
        row += 2
        ws[f'A{row}'] = "📈 수집 통계"
        self._apply_style(ws[f'A{row}'], self.styles['table_name'])
        row += 1
        
        stats = metadata['statistics']
        stat_items = [
            ("총 테이블 수:", f"{stats['total_tables']}개"),
            ("총 컬럼 수:", f"{stats['total_columns']}개"),
            ("총 외래키 수:", f"{stats['total_foreign_keys']}개"),
            ("수집 소요시간:", f"{stats['collection_duration_ms']}ms")
        ]
        
        for label, value in stat_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            self._apply_style(ws[f'A{row}'], {
                'font': Font(name='맑은 고딕', size=10, bold=True),
                'alignment': Alignment(horizontal='right', vertical='center')
            })
            self._apply_style(ws[f'B{row}'], {
                'font': Font(name='맑은 고딕', size=10),
                'alignment': Alignment(horizontal='left', vertical='center')
            })
            row += 1
            
        # 컬럼 너비 조정
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        
    def _create_table_sheet(self, table_name, table_data, foreign_keys):
        """테이블 상세 시트 생성 (이미지 양식에 맞춤)"""
        ws = self.workbook.create_sheet(f"📋 {table_name}")
        
        # 테이블 정보 영역 (A1:J7)
        
        # 1행: 테이블명
        ws['A1'] = "테이블명"
        ws['C1'] = table_name
        ws.merge_cells('A1:B1')
        ws.merge_cells('C1:J1')
        self._apply_style(ws['A1'], self.styles['info_label'])
        self._apply_style(ws['B1'], self.styles['info_label'])  # 병합된 셀도 테두리 적용
        self._apply_style(ws['C1'], self.styles['info_value'])
        # 병합된 나머지 셀들에도 테두리 적용
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J']:
            self._apply_style(ws[f'{col}1'], self.styles['info_value'])
        
        # 2행: 논리명 (빈 칸)
        ws['A2'] = "논리명"
        ws.merge_cells('A2:B2')
        ws.merge_cells('C2:J2')
        self._apply_style(ws['A2'], self.styles['info_label'])
        self._apply_style(ws['B2'], self.styles['info_label'])
        self._apply_style(ws['C2'], self.styles['info_value'])
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J']:
            self._apply_style(ws[f'{col}2'], self.styles['info_value'])
        
        # 3행: 테이블 설명
        ws['A3'] = "테이블 설명"
        ws['C3'] = table_data['table_info']['comment'] or ''
        ws.merge_cells('A3:B3')
        ws.merge_cells('C3:J3')
        self._apply_style(ws['A3'], self.styles['info_label'])
        self._apply_style(ws['B3'], self.styles['info_label'])
        self._apply_style(ws['C3'], self.styles['info_value'])
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J']:
            self._apply_style(ws[f'{col}3'], self.styles['info_value'])
        
        # Primary Key, Foreign Key, Index 정보 수집
        primary_keys = []
        foreign_key_info = []
        
        for col_info in table_data['columns']:
            if col_info['key'] == 'PRI':
                primary_keys.append(col_info['name'])
                
        for fk in foreign_keys:
            if fk['table_name'] == table_name:
                foreign_key_info.append(f"{fk['column_name']}")
        
        # 4행: PRIMARY KEY
        ws['A4'] = "PRIMARY KEY"
        ws['C4'] = ', '.join(primary_keys) if primary_keys else ''
        ws.merge_cells('A4:B4')
        ws.merge_cells('C4:J4')
        self._apply_style(ws['A4'], self.styles['info_label'])
        self._apply_style(ws['B4'], self.styles['info_label'])
        self._apply_style(ws['C4'], self.styles['info_value'])
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J']:
            self._apply_style(ws[f'{col}4'], self.styles['info_value'])
        
        # 5행: FOREIGN KEY
        ws['A5'] = "FOREIGN KEY"
        ws['C5'] = ', '.join(foreign_key_info) if foreign_key_info else ''
        ws.merge_cells('A5:B5')
        ws.merge_cells('C5:J5')
        self._apply_style(ws['A5'], self.styles['info_label'])
        self._apply_style(ws['B5'], self.styles['info_label'])
        self._apply_style(ws['C5'], self.styles['info_value'])
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J']:
            self._apply_style(ws[f'{col}5'], self.styles['info_value'])
        
        # 6행: INDEX
        ws['A6'] = "INDEX"
        ws.merge_cells('A6:B6')
        ws.merge_cells('C6:J6')
        self._apply_style(ws['A6'], self.styles['info_label'])
        self._apply_style(ws['B6'], self.styles['info_label'])
        self._apply_style(ws['C6'], self.styles['info_value'])
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J']:
            self._apply_style(ws[f'{col}6'], self.styles['info_value'])
        
        # 7행: UNIQUE INDEX
        ws['A7'] = "UNIQUE INDEX"
        ws.merge_cells('A7:B7')
        ws.merge_cells('C7:J7')
        self._apply_style(ws['A7'], self.styles['info_label'])
        self._apply_style(ws['B7'], self.styles['info_label'])
        self._apply_style(ws['C7'], self.styles['info_value'])
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J']:
            self._apply_style(ws[f'{col}7'], self.styles['info_value'])
        
        # 8행: 컬럼 헤더 (빈 행 제거)
        headers = ['NO', 'PK', 'AI', 'FK', 'NULL', '컬럼명', 'TYPE', 'DEFAULT', '설명', '참조 테이블']
        row = 8
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_style(cell, self.styles['header'])
            
        # 컬럼 정보 추가 (순차적 NO 부여)
        for idx, col_info in enumerate(table_data['columns'], 1):
            row += 1
            
            # 외래키 참조 정보 찾기 및 FK 여부 판별
            fk_ref = ""
            is_foreign_key = False
            for fk in foreign_keys:
                if (fk['table_name'] == table_name and 
                    fk['column_name'] == col_info['name']):
                    fk_ref = f"{fk['referenced_table_name']}.{fk['referenced_column_name']}"
                    is_foreign_key = True
                    break
                    
            # 데이터 입력
            data_row = [
                idx,  # NO (순차적)
                'Y' if col_info['key'] == 'PRI' else '',  # PK
                'Y' if 'auto_increment' in str(col_info['extra']).lower() else '',  # AI
                'Y' if is_foreign_key else '',  # FK (외래키 정보로 정확히 판별)
                'Y' if col_info['nullable'] in ['YES', 'Y', 1, '1'] else '',  # NULL
                col_info['name'],  # 컬럼명
                col_info['type'],  # TYPE
                col_info['default'] or '',  # DEFAULT
                col_info['comment'] or '',  # 설명
                fk_ref  # 참조 테이블
            ]
            
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col, value=value)
                
                # 스타일 적용
                if col in [1, 2, 3, 4, 5]:  # NO, PK, AI, FK, NULL
                    style = self.styles['data_center']
                    # PK가 Y인 경우 빨간색으로 강조
                    if col == 2 and value == 'Y':
                        style = self.styles['pk']
                else:
                    style = self.styles['data']
                    
                self._apply_style(cell, style)
                
        # 컬럼 너비 조정 (A+B 병합 셀을 고려하여 A컬럼을 넓게 설정)
        column_widths = [12, 4, 6, 6, 8, 20, 20, 15, 30, 20]  # A컬럼을 12로 확대, B컬럼을 4로 축소
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
            
    def _create_foreign_key_sheet(self, foreign_keys):
        """외래키 관계 시트 생성"""
        ws = self.workbook.create_sheet("🔗 외래키 관계")
        
        # 제목
        ws['A1'] = "외래키 관계"
        ws.merge_cells('A1:E1')
        self._apply_style(ws['A1'], self.styles['title'])
        
        # 헤더
        headers = ['테이블명', '컬럼명', '참조 테이블', '참조 컬럼', '제약조건명']
        row = 3
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_style(cell, self.styles['header'])
            
        # 외래키 데이터
        for fk in foreign_keys:
            row += 1
            
            data_row = [
                fk['table_name'],
                fk['column_name'],
                fk['referenced_table_name'],
                fk['referenced_column_name'],
                fk['constraint_name']
            ]
            
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col, value=value)
                self._apply_style(cell, self.styles['data'])
                
        # 컬럼 너비 조정
        column_widths = [20, 20, 20, 20, 25]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
            
    def _format_metadata_for_excel(self, metadata):
        """Excel 생성을 위한 메타데이터 포맷"""
        formatted = {
            'database_info': metadata['connection_info'],
            'tables_by_name': {},
            'foreign_keys': metadata['foreign_keys'],
            'indexes': metadata.get('indexes', []),
            'statistics': metadata['statistics']
        }
        
        # 테이블별로 그룹화
        for row in metadata['tables']:
            table_name = row['table_name']
            if table_name not in formatted['tables_by_name']:
                formatted['tables_by_name'][table_name] = {
                    'table_info': {
                        'name': table_name,
                        'comment': row['table_comment']
                    },
                    'columns': []
                }
                
            formatted['tables_by_name'][table_name]['columns'].append({
                'position': int(row['column_position']) if row['column_position'] else 0,
                'name': row['column_name'],
                'type': row['data_type'],
                'default': row['default_value'],
                'nullable': row['is_nullable'],
                'key': row['key_type'],
                'extra': row['extra'],
                'comment': row['column_comment']
            })
            
        # 컬럼을 position 순으로 정렬
        for table_data in formatted['tables_by_name'].values():
            table_data['columns'].sort(key=lambda x: x['position'] or 0)
            
        return formatted
        
    def _apply_style(self, cell, style_dict):
        """셀에 스타일 적용"""
        if 'font' in style_dict:
            cell.font = style_dict['font']
        if 'alignment' in style_dict:
            cell.alignment = style_dict['alignment']
        if 'fill' in style_dict:
            cell.fill = style_dict['fill']
        if 'border' in style_dict:
            cell.border = style_dict['border']


# 싱글톤 인스턴스
excel_generator = DBSpecExcelGenerator()